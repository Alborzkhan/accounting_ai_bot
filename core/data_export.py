# core/data_export.py
"""خروجی گرفتن از داده‌ها در فرمت‌های Excel و CSV"""

import sys, os, csv, io

from datetime import datetime
from typing import Dict, List, Optional
from sqlalchemy import func
from sqlalchemy.orm import sessionmaker
from database.models import init_db, Account, JournalEntry, JournalLine, Customer, Vendor


class DataExporter:
    """خروجی‌گیری از داده‌های حسابداری."""

    def __init__(self, db_path: str = "accounting.db") -> None:
        self.engine = init_db(db_path)
        self.Session = sessionmaker(bind=self.engine)

    def export_journal_csv(self, user_id: int,
                           date_from: Optional[datetime] = None,
                           date_to: Optional[datetime] = None,
                           dialect: str = "excel") -> str:
        """خروجی CSV از تمام اسناد حسابداری."""
        session = self.Session()
        try:
            q = session.query(
                JournalEntry.id,
                JournalEntry.date,
                JournalEntry.description,
                JournalLine.side,
                JournalLine.amount,
                Account.code,
                Account.name,
            ).join(JournalLine, JournalLine.entry_id == JournalEntry.id
            ).join(Account, Account.id == JournalLine.account_id
            ).filter(JournalEntry.user_id == user_id)

            if date_from:
                q = q.filter(JournalEntry.date >= date_from)
            if date_to:
                q = q.filter(JournalEntry.date <= date_to)

            rows = q.order_by(JournalEntry.date, JournalEntry.id).all()

            output = io.StringIO()
            writer = csv.writer(output, dialect=dialect)
            writer.writerow(["شماره سند", "تاریخ", "شرح", "کد حساب", "نام حساب", "بدهکار", "بستانکار"])

            for r in rows:
                debit = r.amount if r.side == "debit" else 0
                credit = r.amount if r.side == "credit" else 0
                date_str = r.date.strftime("%Y/%m/%d") if r.date else ""
                writer.writerow([r.id, date_str, r.description, r.code, r.name,
                                f"{debit:,.0f}", f"{credit:,.0f}"])

            return output.getvalue()
        finally:
            session.close()

    def export_journal_excel(self, user_id: int,
                             date_from: Optional[datetime] = None,
                             date_to: Optional[datetime] = None) -> bytes:
        """خروجی Excel از اسناد حسابداری با openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "اسناد حسابداری"
        ws.rtl = True
        ws.sheet_view.rightToLeft = True

        # هدر
        header_font = Font(name="Vazirmatn", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A4A6F", end_color="1A4A6F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = ["شماره سند", "تاریخ", "شرح", "کد حساب", "نام حساب", "بدهکار", "بستانکار"]
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # داده‌ها
        session = self.Session()
        try:
            q = session.query(
                JournalEntry.id, JournalEntry.date, JournalEntry.description,
                JournalLine.side, JournalLine.amount,
                Account.code, Account.name,
            ).join(JournalLine, JournalLine.entry_id == JournalEntry.id
            ).join(Account, Account.id == JournalLine.account_id
            ).filter(JournalEntry.user_id == user_id)

            if date_from:
                q = q.filter(JournalEntry.date >= date_from)
            if date_to:
                q = q.filter(JournalEntry.date <= date_to)

            rows = q.order_by(JournalEntry.date, JournalEntry.id).all()

            data_font = Font(name="Vazirmatn", size=10)
            num_alignment = Alignment(horizontal="left", vertical="center")
            center_alignment = Alignment(horizontal="center", vertical="center")

            for idx, r in enumerate(rows, 2):
                debit = r.amount if r.side == "debit" else 0
                credit = r.amount if r.side == "credit" else 0
                date_str = r.date.strftime("%Y/%m/%d") if r.date else ""

                ws.cell(row=idx, column=1, value=r.id).font = data_font
                ws.cell(row=idx, column=2, value=date_str).font = data_font
                ws.cell(row=idx, column=3, value=r.description).font = data_font
                ws.cell(row=idx, column=4, value=r.code).font = data_font
                ws.cell(row=idx, column=5, value=r.name).font = data_font
                ws.cell(row=idx, column=6, value=debit).font = data_font
                ws.cell(row=idx, column=7, value=credit).font = data_font

                for col in range(1, 8):
                    ws.cell(row=idx, column=col).border = thin_border
                    ws.cell(row=idx, column=col).alignment = center_alignment if col <= 5 else num_alignment
                ws.cell(row=idx, column=3).alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)

            # تنظیم عرض ستون‌ها
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 18

        finally:
            session.close()

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_customers_csv(self, user_id: int) -> str:
        """خروجی CSV از لیست مشتریان."""
        session = self.Session()
        try:
            customers = session.query(Customer).filter(
                (Customer.user_id == user_id) | (Customer.user_id.is_(None))
            ).all()

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["کد", "نام", "موبایل", "تلفن", "کد ملی", "کد اقتصادی", "آدرس"])

            for c in customers:
                writer.writerow([c.id, c.name, c.mobile, c.phone,
                                c.national_id, c.economic_code, c.address])

            return output.getvalue()
        finally:
            session.close()

    def export_vendors_csv(self, user_id: int) -> str:
        """خروجی CSV از لیست فروشندگان."""
        session = self.Session()
        try:
            vendors = session.query(Vendor).filter(
                (Vendor.user_id == user_id) | (Vendor.user_id.is_(None))
            ).all()

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["کد", "نام", "موبایل", "تلفن", "کد ملی", "کد اقتصادی"])

            for v in vendors:
                writer.writerow([v.id, v.name, v.mobile, v.phone,
                                v.national_id, v.economic_code])

            return output.getvalue()
        finally:
            session.close()

    def export_trial_balance_csv(self, user_id: int) -> str:
        """خروجی CSV از تراز آزمایشی."""
        from core.accounting_engine import AccountingEngine
        engine = AccountingEngine()
        balances = engine.get_trial_balance(user_id=user_id)

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["کد", "نام حساب", "بدهکار", "بستانکار", "نوع"])

        for row in balances:
            if row.total_debit != 0 or row.total_credit != 0:
                writer.writerow([
                    row.code, row.name,
                    f"{row.total_debit:,.0f}",
                    f"{row.total_credit:,.0f}",
                    getattr(row, 'type', '') or '',
                ])

        return output.getvalue()
